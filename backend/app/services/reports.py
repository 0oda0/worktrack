"""Сводка и рейтинг — поверх чистого compute_stats."""

from datetime import date, timedelta

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import ROLE_WORKER, AttendanceRecord, Holiday, User
from app.services.settings_store import TZ
from app.services.timesheet import compute_stats, is_weekend_or_holiday

WORK_START_HOUR = 9  # позже — опоздание
# веса рейтинга
W_OVERTIME = 2.0
W_HOURS = 0.1
W_LATENESS = 3.0


def _holidays(db: Session) -> set[date]:
    return {h.date for h in db.scalars(select(Holiday)).all()}


def _records_for(db: Session, user_id: int, start: date, end: date):
    return db.scalars(
        select(AttendanceRecord).where(
            AttendanceRecord.user_id == user_id,
            AttendanceRecord.work_date >= start,
            AttendanceRecord.work_date <= end,
        )
    ).all()


def summary(db: Session, users: list[User], start: date, end: date) -> list[dict]:
    holidays = _holidays(db)
    rows = []
    for u in users:
        recs = _records_for(db, u.id, start, end)
        stats = compute_stats(
            [(r.work_date, r.check_in, r.check_out) for r in recs],
            holidays,
            u.hire_date,
            start,
            end,
        )
        rows.append(
            {
                "user_id": u.id,
                "full_name": u.full_name,
                "audience": u.audience,
                "hire_date": u.hire_date,
                **stats.__dict__,
            }
        )
    return rows


def rating(db: Session, users: list[User], start: date, end: date) -> list[dict]:
    holidays = _holidays(db)
    rows = []
    for u in users:
        recs = _records_for(db, u.id, start, end)
        stats = compute_stats(
            [(r.work_date, r.check_in, r.check_out) for r in recs], holidays, u.hire_date, start, end
        )
        lateness = 0
        for r in recs:
            if is_weekend_or_holiday(r.work_date, holidays):
                continue
            local_in = r.check_in.astimezone(TZ)
            if (local_in.hour, local_in.minute) > (WORK_START_HOUR, 0):
                lateness += 1
        score = (
            stats.overtime * W_OVERTIME
            + stats.total_hours * W_HOURS
            - lateness * W_LATENESS
        )
        rows.append(
            {
                "user_id": u.id,
                "full_name": u.full_name,
                "audience": u.audience,
                "total_hours": stats.total_hours,
                "overtime": stats.overtime,
                "weekend_hours": stats.weekend_hours,
                "lateness": lateness,
                "score": round(score, 2),
            }
        )
    rows.sort(key=lambda r: r["score"], reverse=True)
    return rows


# ---------- экспорт: детализация по сессиям ----------

REPORT_HEADERS = [
    "ФИО",
    "Telegram Username",  # у нас здесь email (телеграма нет)
    "Сектор",  # у нас — аудитория
    "Начало сессии",
    "Конец сессии",
    "Длительность сессии",
]
_COL_WIDTHS = {"A": 40, "B": 20, "C": 8, "D": 21, "E": 21, "F": 22}
_DURATION_FMT = "[h]:mm:ss"
_DT_FMT = "%Y-%m-%d %H:%M:%S"


def session_report(db: Session, users: list[User], start: date, end: date) -> list[dict]:
    """По каждому сотруднику с сессиями в периоде: список смен (по убыванию даты) + сумма длительностей."""
    report = []
    for u in users:
        recs = sorted(_records_for(db, u.id, start, end), key=lambda r: r.check_in, reverse=True)
        if not recs:
            continue  # без сессий человека не показываем (как в эталоне)
        sessions = []
        total = timedelta()
        for r in recs:
            dur = (r.check_out - r.check_in) if r.check_out else None
            if dur:
                total += dur
            sessions.append(
                {
                    "start": r.check_in.astimezone(TZ).strftime(_DT_FMT),
                    "end": r.check_out.astimezone(TZ).strftime(_DT_FMT) if r.check_out else None,
                    "duration": dur,
                }
            )
        report.append(
            {
                "full_name": u.full_name,
                "email": u.email,
                "audience": u.audience,
                "sessions": sessions,
                "total": total,
            }
        )
    return report


def export_workbook(report: list[dict], sheet_title: str = "Отчёт") -> Workbook:
    """Строит книгу в формате эталона: строка на сессию, «Итого, ФИО», пустая строка-разделитель."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(REPORT_HEADERS)
    for person in report:
        for s in person["sessions"]:
            ws.append(
                [person["full_name"], person["email"], person["audience"], s["start"], s["end"], s["duration"]]
            )
            if s["duration"] is not None:
                ws.cell(ws.max_row, 6).number_format = _DURATION_FMT
        ws.append([f"Итого, {person['full_name']}", None, None, None, None, person["total"]])
        ws.cell(ws.max_row, 6).number_format = _DURATION_FMT
        ws.append([])  # разделитель между сотрудниками
    for col, width in _COL_WIDTHS.items():
        ws.column_dimensions[col].width = width
    return wb


def workers_in_scope(db: Session, audience: str | None = None) -> list[User]:
    """Работники (admin и leader видят всех); audience — опциональный фильтр."""
    q = select(User).where(User.role == ROLE_WORKER, User.is_active.is_(True))
    if audience:
        q = q.where(User.audience == audience)
    return list(db.scalars(q.order_by(User.full_name)).all())

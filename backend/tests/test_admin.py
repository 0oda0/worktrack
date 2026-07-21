"""Спринт B4 — админка: пользователи, отметки, now-working, отчёты, рейтинг, Excel, настройки."""

from datetime import date, timedelta
from io import BytesIO

from openpyxl import load_workbook

IN_ZONE = {"lat": 55.7552, "lng": 37.7148}  # внутри полигона МТУСИ


def register(client, email, name="U"):
    return client.post(
        "/api/auth/register", json={"full_name": name, "email": email, "password": "secret1"}
    )


def login(client, email):
    return client.post(
        "/api/auth/login", json={"email": email, "password": "secret1"}
    ).json()["token"]


def auth(t):
    return {"Authorization": f"Bearer {t}"}


def admin_token(client):
    return client.post(
        "/api/auth/login", json={"email": "admin@worktrack.ru", "password": "admin"}
    ).json()["token"]


def approve_user(client, email, role="worker", audience="203", hire="2026-01-01"):
    uid = register(client, email).json()["id"]
    client.post(
        f"/api/admin/users/{uid}/approve",
        headers=auth(admin_token(client)),
        json={"role": role, "audience": audience, "hire_date": hire},
    )
    return uid, login(client, email)


def past(days=3):
    """Прошедший рабочий день: опоздания и переработки в выходные не считаются,
    поэтому дата не должна «уезжать» на субботу в зависимости от дня запуска."""
    d = date.today() - timedelta(days=days)
    while d.weekday() >= 5:
        d -= timedelta(days=1)
    return d.isoformat()


def manual_record(client, tok, atok, pd, ci="09:00:00", co="17:00:00"):
    req = client.post(
        "/api/requests", headers=auth(tok),
        json={"work_date": pd, "check_in": f"{pd}T{ci}", "check_out": f"{pd}T{co}"},
    ).json()
    client.post(f"/api/requests/{req['id']}/approve", headers=auth(atok), json={})


# ---- пользователи ----

def test_list_and_filter_users(client):
    approve_user(client, "a@mtuci.ru", audience="203")
    approve_user(client, "b@mtuci.ru", audience="903")
    atok = admin_token(client)
    all_users = client.get("/api/admin/users", headers=auth(atok)).json()
    assert len(all_users) >= 3  # admin + 2
    only203 = client.get("/api/admin/users?audience=203", headers=auth(atok)).json()
    assert {u["audience"] for u in only203} == {"203"}


def test_patch_user_role_and_audience(client):
    uid, _ = approve_user(client, "w@mtuci.ru")
    atok = admin_token(client)
    r = client.patch(
        f"/api/admin/users/{uid}", headers=auth(atok),
        json={"role": "leader", "audience": "906"},
    )
    assert r.status_code == 200
    assert r.json()["role"] == "leader" and r.json()["audience"] == "906"


def test_deactivate_blocks_login(client):
    uid, _ = approve_user(client, "w@mtuci.ru")
    atok = admin_token(client)
    client.patch(f"/api/admin/users/{uid}", headers=auth(atok), json={"is_active": False})
    r = client.post("/api/auth/login", json={"email": "w@mtuci.ru", "password": "secret1"})
    assert r.status_code == 403


def test_worker_cannot_list_users(client):
    _, tok = approve_user(client, "w@mtuci.ru")
    assert client.get("/api/admin/users", headers=auth(tok)).status_code == 403


# ---- отметки: правка/удаление + скоуп ----

def test_admin_edits_record_changes_timesheet(client):
    uid, tok = approve_user(client, "w@mtuci.ru")
    atok = admin_token(client)
    pd = past()
    manual_record(client, tok, atok, pd, "09:00:00", "17:00:00")  # 8ч
    rec = client.get(f"/api/admin/records?user_id={uid}", headers=auth(atok)).json()[0]

    client.patch(
        f"/api/admin/records/{rec['id']}", headers=auth(atok),
        json={"check_out": f"{pd}T19:00:00"},  # теперь 10ч
    )
    ts = client.get(f"/api/attendance/timesheet?start={pd}&end={pd}", headers=auth(tok)).json()
    assert ts["days"][0]["hours"] == 10.0
    assert ts["stats"]["overtime"] == 1.0  # 10-9


def test_edit_record_checkout_before_checkin_rejected(client):
    uid, tok = approve_user(client, "w@mtuci.ru")
    atok = admin_token(client)
    pd = past()
    manual_record(client, tok, atok, pd)
    rec = client.get(f"/api/admin/records?user_id={uid}", headers=auth(atok)).json()[0]
    r = client.patch(
        f"/api/admin/records/{rec['id']}", headers=auth(atok),
        json={"check_out": f"{pd}T08:00:00"},
    )
    assert r.status_code == 422


def test_delete_record(client):
    uid, tok = approve_user(client, "w@mtuci.ru")
    atok = admin_token(client)
    pd = past()
    manual_record(client, tok, atok, pd)
    rec = client.get(f"/api/admin/records?user_id={uid}", headers=auth(atok)).json()[0]
    assert client.delete(f"/api/admin/records/{rec['id']}", headers=auth(atok)).status_code == 204
    assert client.get(f"/api/admin/records?user_id={uid}", headers=auth(atok)).json() == []


def test_leader_sees_all_audiences_and_filter(client):
    uid_a, tok_a = approve_user(client, "a@mtuci.ru", audience="203")
    uid_b, tok_b = approve_user(client, "b@mtuci.ru", audience="903")
    _, ltok = approve_user(client, "lead@mtuci.ru", role="leader", audience="203")
    atok = admin_token(client)
    pd = past()
    manual_record(client, tok_a, atok, pd)
    manual_record(client, tok_b, atok, pd)
    seen = client.get("/api/admin/records", headers=auth(ltok)).json()
    assert {r["audience"] for r in seen} == {"203", "903"}  # лидер видит всех
    filtered = client.get("/api/admin/records?audience=903", headers=auth(ltok)).json()
    assert {r["audience"] for r in filtered} == {"903"}


def test_leader_can_delete_any_audience_record(client):
    uid_b, tok_b = approve_user(client, "b@mtuci.ru", audience="903")
    _, ltok = approve_user(client, "lead@mtuci.ru", role="leader", audience="203")
    atok = admin_token(client)
    pd = past()
    manual_record(client, tok_b, atok, pd)
    rec = client.get("/api/admin/records", headers=auth(atok)).json()[0]
    assert client.delete(f"/api/admin/records/{rec['id']}", headers=auth(ltok)).status_code == 204


# ---- now-working ----

def test_now_working(client):
    _, tok = approve_user(client, "w@mtuci.ru")
    atok = admin_token(client)
    client.post("/api/attendance/check-in", headers=auth(tok), json=IN_ZONE)
    nw = client.get("/api/admin/now-working", headers=auth(atok)).json()
    assert len(nw) == 1 and nw[0]["full_name"] == "U"


def test_now_working_audience_filter(client):
    _, tok_a = approve_user(client, "a@mtuci.ru", audience="203")
    _, tok_b = approve_user(client, "b@mtuci.ru", audience="903")
    atok = admin_token(client)
    client.post("/api/attendance/check-in", headers=auth(tok_a), json=IN_ZONE)
    client.post("/api/attendance/check-in", headers=auth(tok_b), json=IN_ZONE)
    nw = client.get("/api/admin/now-working?audience=903", headers=auth(atok)).json()
    assert {r["audience"] for r in nw} == {"903"}


# ---- отчёты / рейтинг / Excel ----

def test_summary_matches_timesheet(client):
    uid, tok = approve_user(client, "w@mtuci.ru")
    atok = admin_token(client)
    pd = past()
    manual_record(client, tok, atok, pd, "09:00:00", "17:00:00")
    s = client.get(f"/api/reports/summary?start={pd}&end={pd}", headers=auth(atok)).json()
    row = next(r for r in s if r["user_id"] == uid)
    assert row["total_hours"] == 8.0


def test_reports_audience_filter(client):
    approve_user(client, "a@mtuci.ru", audience="203")
    approve_user(client, "b@mtuci.ru", audience="903")
    atok = admin_token(client)
    s = client.get("/api/reports/summary?audience=903", headers=auth(atok)).json()
    assert {r["audience"] for r in s} == {"903"}
    r = client.get("/api/reports/rating?audience=903", headers=auth(atok)).json()
    assert {x["audience"] for x in r} == {"903"}


def test_rating_sorted_by_score(client):
    _, tok1 = approve_user(client, "hard@mtuci.ru")
    uid2, tok2 = approve_user(client, "late@mtuci.ru")
    atok = admin_token(client)
    pd = past()
    manual_record(client, tok1, atok, pd, "09:00:00", "20:00:00")  # переработка
    manual_record(client, tok2, atok, pd, "11:00:00", "13:00:00")  # опоздание, мало часов
    r = client.get(f"/api/reports/rating?start={pd}&end={pd}", headers=auth(atok)).json()
    scores = [x["score"] for x in r]
    assert scores == sorted(scores, reverse=True)
    assert r[0]["full_name"] == "U"  # оба "U", но hard-worker выше
    late = next(x for x in r if x["user_id"] == uid2)
    assert late["lateness"] == 1


def test_export_xlsx(client):
    from datetime import timedelta

    uid, tok = approve_user(client, "w@mtuci.ru")
    atok = admin_token(client)
    pd = past()
    manual_record(client, tok, atok, pd, "09:00:00", "17:00:00")
    resp = client.get(f"/api/reports/export.xlsx?start={pd}&end={pd}", headers=auth(atok))
    assert resp.status_code == 200
    assert "spreadsheet" in resp.headers["content-type"]
    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    # детализация по сессиям: заголовки + строка сессии + «Итого»
    assert ws.cell(1, 1).value == "ФИО"
    assert ws.cell(1, 6).value == "Длительность сессии"
    assert ws.cell(2, 4).value == f"{pd} 09:00:00"  # начало сессии
    assert ws.cell(2, 6).value == timedelta(hours=8)  # длительность
    assert ws.cell(3, 1).value.startswith("Итого,")


def test_worker_cannot_access_reports(client):
    _, tok = approve_user(client, "w@mtuci.ru")
    assert client.get("/api/reports/summary", headers=auth(tok)).status_code == 403
    assert client.get("/api/reports/rating", headers=auth(tok)).status_code == 403


# ---- настройки геозоны ----

def test_get_and_update_settings_affects_zone(client):
    atok = admin_token(client)
    s = client.get("/api/admin/settings", headers=auth(atok)).json()
    assert s["office_radius_m"] == 200

    # сузим радиус до 10 м и сместим офис далеко → прежняя точка станет вне зоны
    client.put(
        "/api/admin/settings", headers=auth(atok),
        json={"office_lat": 10.0, "office_lng": 10.0, "office_radius_m": 10},
    )
    _, tok = approve_user(client, "w@mtuci.ru")
    r = client.post("/api/attendance/check-in", headers=auth(tok), json=IN_ZONE)
    assert r.json()["out_of_zone_in"] is True


def test_now_working_skips_stale_open_shift(client, db_session):
    """Незакрытая смена за прошлый день не должна числиться «сейчас на работе»."""
    from datetime import timedelta as td
    from app.models import AttendanceRecord

    _, tok = approve_user(client, "stale@mtuci.ru")
    atok = admin_token(client)
    client.post("/api/attendance/check-in", headers=auth(tok), json=IN_ZONE)

    rec = db_session.query(AttendanceRecord).one()
    rec.work_date = rec.work_date - td(days=1)
    rec.check_in = rec.check_in - td(days=1)
    db_session.commit()

    nw = client.get("/api/admin/now-working", headers=auth(atok)).json()
    assert nw == []
